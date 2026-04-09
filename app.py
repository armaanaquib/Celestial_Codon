"""
Celestial Codon — DNA Analysis Platform
========================================
A Flask-based bioinformatics web application for analyzing DNA sequences.

Features:
  - GC content, melting temperature, nucleotide frequency
  - 6-frame ORF detection with canvas visualization
  - Protein properties: MW, pI, hydrophobicity (Kyte-Doolittle)
  - Codon bias comparison (E. coli, Human, Yeast reference tables)
  - Restriction enzyme site detection (30+ enzymes, N-wildcard support)
  - Motif library scan (TATA-box, Kozak, NF-kB, SP1, AP1, etc.)
  - SNP / point mutation simulator with disease lookup (MyVariant.info)
  - Primer design, restriction digest simulator, PCR product predictor
  - CpG island detection, secondary structure prediction (Chou-Fasman)
  - AI assistant (Armaani) powered by Groq API (llama-3.1-8b-instant)
  - Export to Excel (.xlsx), FASTA, and JSON

Setup:
  pip install -r requirements.txt
  set GROQ_API_KEY=your_key_here   # get free key at https://console.groq.com
  python app.py

Author: Armaan Aquib
License: armaani
"""

from flask import Flask, render_template, request, jsonify
from collections import Counter
import base64, io, math
import os
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend — required for server-side chart generation
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

# ── Gemini API (DISABLED — requires paid subscription / quota exceeded) ──────
# Uncomment below if you have a valid Gemini API key with available quota.
# from google import genai
# gemini_client = genai.Client(api_key=os.getenv("GEMINI_API_KEY", "YOUR_KEY_HERE"))
# NOTE: Free tier quota is very limited. Use Groq instead.

# ── Groq API (FREE TIER — active) ────────────────────────────────────────────
try:
    from groq import Groq as GroqClient
    _groq_available = True
    print("Groq package found. Client will be created per-request using GROQ_API_KEY env var.")
except ImportError:
    _groq_available = False
    print("WARNING: groq package not installed. Run: pip install groq")

def get_groq_client():
    """Create Groq client fresh each call so it always picks up the current env var."""
    if not _groq_available:
        return None
    key = os.getenv("GROQ_API_KEY", "")
    if not key:
        return None
    return GroqClient(api_key=key)

app = Flask(__name__)

# ── Tables ────────────────────────────────────────────────────────────────────
CODON_TABLE = {
    'UUU':'Phe','UUC':'Phe','UUA':'Leu','UUG':'Leu',
    'CUU':'Leu','CUC':'Leu','CUA':'Leu','CUG':'Leu',
    'AUU':'Ile','AUC':'Ile','AUA':'Ile','AUG':'Met',
    'GUU':'Val','GUC':'Val','GUA':'Val','GUG':'Val',
    'UCU':'Ser','UCC':'Ser','UCA':'Ser','UCG':'Ser',
    'CCU':'Pro','CCC':'Pro','CCA':'Pro','CCG':'Pro',
    'ACU':'Thr','ACC':'Thr','ACA':'Thr','ACG':'Thr',
    'GCU':'Ala','GCC':'Ala','GCA':'Ala','GCG':'Ala',
    'UAU':'Tyr','UAC':'Tyr','UAA':'STOP','UAG':'STOP',
    'CAU':'His','CAC':'His','CAA':'Gln','CAG':'Gln',
    'AAU':'Asn','AAC':'Asn','AAA':'Lys','AAG':'Lys',
    'GAU':'Asp','GAC':'Asp','GAA':'Glu','GAG':'Glu',
    'UGU':'Cys','UGC':'Cys','UGA':'STOP','UGG':'Trp',
    'CGU':'Arg','CGC':'Arg','CGA':'Arg','CGG':'Arg',
    'AGU':'Ser','AGC':'Ser','AGA':'Arg','AGG':'Arg',
    'GGU':'Gly','GGC':'Gly','GGA':'Gly','GGG':'Gly',
}

# Amino acid molecular weights (Da), pKa values, hydrophobicity (Kyte-Doolittle)
AA_MW = {'Ala':89.09,'Arg':174.20,'Asn':132.12,'Asp':133.10,'Cys':121.16,
         'Gln':146.15,'Glu':147.13,'Gly':75.03,'His':155.16,'Ile':131.17,
         'Leu':131.17,'Lys':146.19,'Met':149.21,'Phe':165.19,'Pro':115.13,
         'Ser':105.09,'Thr':119.12,'Trp':204.23,'Tyr':181.19,'Val':117.15}
AA_PKA = {'Asp':3.9,'Glu':4.1,'His':6.0,'Cys':8.3,'Tyr':10.1,'Lys':10.5,'Arg':12.5}
AA_HYDRO = {'Ile':4.5,'Val':4.2,'Leu':3.8,'Phe':2.8,'Cys':2.5,'Met':1.9,'Ala':1.8,
            'Gly':-0.4,'Thr':-0.7,'Ser':-0.8,'Trp':-0.9,'Tyr':-1.3,'Pro':-1.6,
            'His':-3.2,'Glu':-3.5,'Gln':-3.5,'Asp':-3.5,'Asn':-3.5,'Lys':-3.9,'Arg':-4.5}

RESTRICTION_ENZYMES = {
    'EcoRI':'GAATTC','BamHI':'GGATCC','HindIII':'AAGCTT','NotI':'GCGGCCGC',
    'XhoI':'CTCGAG','NcoI':'CCATGG','SalI':'GTCGAC','PstI':'CTGCAG',
    'SmaI':'CCCGGG','KpnI':'GGTACC','SacI':'GAGCTC','ApaI':'GGGCCC',
    'SpeI':'ACTAGT','NheI':'GCTAGC','EcoRV':'GATATC','DraI':'TTTAAA',
    'BglII':'AGATCT','ClaI':'ATCGAT','MluI':'ACGCGT','XbaI':'TCTAGA',
    'HpaI':'GTTAAC','ScaI':'AGTACT','StuI':'AGGCCT','SphI':'GCATGC',
    'AgeI':'ACCGGT','AvrII':'CCTAGG','PacI':'TTAATTAA','AscI':'GGCGCGCC',
    'FseI':'GGCCGGCC',
}

# Codon usage frequency tables (per 1000 codons) - E.coli K12, Human, Yeast
CODON_BIAS = {
    'ecoli': {
        'UUU':22.0,'UUC':16.5,'UUA':13.9,'UUG':13.1,'CUU':11.0,'CUC':10.9,'CUA':3.9,'CUG':52.8,
        'AUU':30.3,'AUC':25.3,'AUA':7.6,'AUG':27.9,'GUU':18.5,'GUC':15.3,'GUA':10.9,'GUG':26.1,
        'UCU':8.5,'UCC':8.6,'UCA':7.2,'UCG':8.9,'CCU':7.0,'CCC':5.5,'CCA':8.4,'CCG':23.3,
        'ACU':9.0,'ACC':23.4,'ACA':7.6,'ACG':14.5,'GCU':15.3,'GCC':25.5,'GCA':21.1,'GCG':33.6,
        'UAU':16.2,'UAC':12.2,'UAA':2.0,'UAG':0.3,'CAU':13.0,'CAC':9.7,'CAA':15.3,'CAG':28.7,
        'AAU':21.7,'AAC':21.7,'AAA':33.6,'AAG':10.3,'GAU':32.2,'GAC':19.1,'GAA':39.6,'GAG':17.8,
        'UGU':5.2,'UGC':6.5,'UGA':1.0,'UGG':15.2,'CGU':20.9,'CGC':21.5,'CGA':3.6,'CGG':5.4,
        'AGU':9.1,'AGC':16.1,'AGA':2.1,'AGG':1.2,'GGU':24.7,'GGC':29.8,'GGA':8.0,'GGG':11.1,
    },
    'human': {
        'UUU':17.6,'UUC':20.3,'UUA':7.7,'UUG':12.9,'CUU':13.2,'CUC':19.6,'CUA':7.2,'CUG':39.6,
        'AUU':15.9,'AUC':20.8,'AUA':7.5,'AUG':22.0,'GUU':11.0,'GUC':14.5,'GUA':7.1,'GUG':28.1,
        'UCU':15.2,'UCC':17.7,'UCA':12.2,'UCG':4.4,'CCU':17.5,'CCC':19.8,'CCA':16.9,'CCG':6.9,
        'ACU':13.1,'ACC':18.9,'ACA':15.1,'ACG':6.1,'GCU':18.4,'GCC':27.7,'GCA':15.8,'GCG':7.4,
        'UAU':12.2,'UAC':15.3,'UAA':1.0,'UAG':0.8,'CAU':10.9,'CAC':15.1,'CAA':12.3,'CAG':34.2,
        'AAU':17.0,'AAC':19.1,'AAA':24.4,'AAG':31.9,'GAU':21.8,'GAC':25.1,'GAA':29.0,'GAG':39.6,
        'UGU':10.6,'UGC':12.6,'UGA':1.6,'UGG':13.2,'CGU':4.5,'CGC':10.4,'CGA':6.2,'CGG':11.4,
        'AGU':15.2,'AGC':19.5,'AGA':11.5,'AGG':11.4,'GGU':10.8,'GGC':22.2,'GGA':16.5,'GGG':16.5,
    },
    'yeast': {
        'UUU':26.1,'UUC':18.4,'UUA':26.2,'UUG':27.2,'CUU':12.3,'CUC':5.4,'CUA':13.4,'CUG':10.5,
        'AUU':30.1,'AUC':17.2,'AUA':17.8,'AUG':20.9,'GUU':22.1,'GUC':11.8,'GUA':11.8,'GUG':10.8,
        'UCU':23.5,'UCC':14.2,'UCA':18.7,'UCG':8.6,'CCU':13.5,'CCC':6.8,'CCA':18.3,'CCG':5.3,
        'ACU':20.3,'ACC':12.8,'ACA':17.9,'ACG':8.0,'GCU':21.1,'GCC':12.6,'GCA':16.2,'GCG':6.2,
        'UAU':18.8,'UAC':14.8,'UAA':1.1,'UAG':0.5,'CAU':13.6,'CAC':7.8,'CAA':27.3,'CAG':12.1,
        'AAU':36.1,'AAC':24.8,'AAA':41.9,'AAG':30.8,'GAU':37.6,'GAC':20.2,'GAA':45.5,'GAG':19.2,
        'UGU':8.1,'UGC':4.8,'UGA':0.7,'UGG':10.4,'CGU':6.4,'CGC':2.6,'CGA':3.0,'CGG':1.7,
        'AGU':14.2,'AGC':9.8,'AGA':21.3,'AGG':9.2,'GGU':23.9,'GGC':9.8,'GGA':10.9,'GGG':6.0,
    }
}

# Known promoter / TF binding motifs
MOTIF_LIBRARY = {
    'TATA-box':       'TATAAA',
    'Kozak':          'GCCACC',
    'Shine-Dalgarno': 'AGGAGG',
    'CpG-island':     'CGCGCG',
    'E-box':          'CANNTG',
    'SP1-site':       'GGGCGG',
    'NF-kB':          'GGGRNNTCC',
    'CAAT-box':       'CCAAT',
    'GC-box':         'GGGCGG',
    'AP1-site':       'TGASTCA',
}

# ── Core helpers ──────────────────────────────────────────────────────────────
def validate(seq):
    seq = seq.upper().strip().replace(' ','').replace('\n','')
    invalid = set(seq) - set('ATCG')
    if invalid:
        raise ValueError(f"Invalid characters: {invalid}")
    if len(seq) < 3:
        raise ValueError("Sequence too short (min 3 bases).")
    return seq

def gc_content(seq):
    return round((seq.count('G') + seq.count('C')) / len(seq) * 100, 2)

def reverse_complement(seq):
    return seq.translate(str.maketrans('ATCG','TAGC'))[::-1]

def transcribe(seq):
    return seq.replace('T','U')

def translate_seq(rna, start=0):
    protein = []
    for i in range(start, len(rna)-2, 3):
        codon = rna[i:i+3]
        if len(codon) < 3: break
        aa = CODON_TABLE.get(codon,'???')
        if aa == 'STOP': break
        protein.append(aa)
    return protein

def melting_temp(seq):
    a,t,g,c = seq.count('A'),seq.count('T'),seq.count('G'),seq.count('C')
    n = len(seq)
    if n < 14:
        return round(2*(a+t)+4*(g+c),2), "Wallace rule"
    return round(81.5+0.41*((g+c)/n*100)-675/n,2), "Marmur formula"

def find_motif_positions(seq, motif):
    """Find all 1-based positions of motif in seq. Supports N wildcard."""
    if 'N' not in motif:
        positions, start = [], 0
        while True:
            pos = seq.find(motif, start)
            if pos == -1: break
            positions.append(pos + 1)
            start = pos + 1
        return positions
    # N-wildcard: convert to regex
    import re as _re
    pattern = _re.compile(motif.replace('N', '.'))
    return [m.start() + 1 for m in pattern.finditer(seq)]

def restriction_sites(seq):
    results = {}
    for enzyme, site in RESTRICTION_ENZYMES.items():
        pos = find_motif_positions(seq, site)
        if pos:
            results[enzyme] = {'site': site, 'positions': pos}
    return results

def palindromes(seq):
    result = []
    for length in range(4, min(12, len(seq))+1, 2):
        for i in range(len(seq)-length+1):
            sub = seq[i:i+length]
            if sub == reverse_complement(sub):
                result.append({'start':i+1,'end':i+length,'seq':sub})
    return result[:12]

def codon_usage(seq):
    rna = transcribe(seq)
    codons = [rna[i:i+3] for i in range(0,len(rna)-2,3) if len(rna[i:i+3])==3]
    counts = Counter(codons)
    return [{'codon':c,'aa':CODON_TABLE.get(c,'???'),'count':n} for c,n in sorted(counts.items())]

# ── New: 6-frame ORF finder ───────────────────────────────────────────────────
def find_all_orfs_6frame(seq):
    stops = {'TAA','TAG','TGA'}
    frames = []
    strands = [('forward', seq), ('reverse', reverse_complement(seq))]
    for strand_name, s in strands:
        for frame in range(3):
            orfs = []
            i = frame
            while i < len(s)-2:
                if s[i:i+3] == 'ATG':
                    for j in range(i, len(s)-2, 3):
                        if s[j:j+3] in stops:
                            length = j+3-i
                            orfs.append({'start':i+1,'end':j+3,'length':length,
                                         'seq':s[i:j+3],'frame':frame+1,'strand':strand_name})
                            break
                i += 3
            frames.append({'strand':strand_name,'frame':frame+1,'orfs':orfs})
    return frames

# ── New: Protein properties ───────────────────────────────────────────────────
def protein_properties(protein_list):
    if not protein_list:
        return None
    mw = sum(AA_MW.get(aa, 110) for aa in protein_list) - 18.02*(len(protein_list)-1)
    # pI estimation via charge curve
    def net_charge(ph):
        charge = 0.0
        for aa in protein_list:
            if aa in AA_PKA:
                pka = AA_PKA[aa]
                if aa in ('Asp','Glu','Cys','Tyr'):
                    charge -= 1/(1+10**(ph-pka))
                else:
                    charge += 1/(1+10**(pka-ph))
        charge += 1/(1+10**(ph-8.0))   # N-terminus
        charge -= 1/(1+10**(3.1-ph))   # C-terminus
        return charge
    lo, hi = 0.0, 14.0
    for _ in range(200):
        mid = (lo+hi)/2
        c = net_charge(mid)
        if abs(c) < 0.0001: break
        if c > 0: lo = mid
        else: hi = mid
    pi = round((lo+hi)/2, 2)
    hydro_vals = [AA_HYDRO.get(aa, 0) for aa in protein_list]
    avg_hydro = round(sum(hydro_vals)/len(hydro_vals), 3)
    # sliding window hydrophobicity (window=9)
    win = 9
    hydro_profile = []
    for i in range(len(hydro_vals)-win+1):
        hydro_profile.append(round(sum(hydro_vals[i:i+win])/win, 3))
    return {
        'mw': round(mw, 2),
        'pi': pi,
        'avg_hydrophobicity': avg_hydro,
        'hydro_profile': hydro_profile,
        'length': len(protein_list),
        'classification': 'Hydrophobic' if avg_hydro > 0 else 'Hydrophilic'
    }

# ── New: Codon bias comparison ────────────────────────────────────────────────
def codon_bias_comparison(seq, organism='ecoli'):
    ref = CODON_BIAS.get(organism, CODON_BIAS['ecoli'])
    rna = transcribe(seq)
    codons = [rna[i:i+3] for i in range(0,len(rna)-2,3) if len(rna[i:i+3])==3]
    total = len(codons)
    if total == 0: return []
    counts = Counter(codons)
    result = []
    for codon, count in sorted(counts.items()):
        seq_freq = round(count/total*1000, 1)
        ref_freq = ref.get(codon, 0)
        result.append({
            'codon': codon,
            'aa': CODON_TABLE.get(codon,'???'),
            'seq_freq': seq_freq,
            'ref_freq': ref_freq,
            'diff': round(seq_freq - ref_freq, 1)
        })
    return result

# ── New: Motif library scan ───────────────────────────────────────────────────
def scan_motif_library(seq):
    results = []
    for name, pattern in MOTIF_LIBRARY.items():
        # Handle degenerate bases (N=any, R=A/G, S=C/G, W=A/T, Y=C/T)
        import re
        regex = pattern.replace('N','[ATCG]').replace('R','[AG]').replace('Y','[CT]') \
                       .replace('S','[GC]').replace('W','[AT]').replace('K','[GT]') \
                       .replace('M','[AC]')
        positions = [m.start()+1 for m in re.finditer(f'(?={regex})', seq)]
        if positions:
            results.append({'name':name,'pattern':pattern,'positions':positions,'count':len(positions)})
    return results

# ── New: SNP / Mutation simulator ────────────────────────────────────────────
def simulate_mutation(seq, position, new_base):
    pos = int(position) - 1
    if pos < 0 or pos >= len(seq):
        raise ValueError("Position out of range.")
    new_base = new_base.upper()
    if new_base not in 'ATCG':
        raise ValueError("Invalid base.")
    mutated = seq[:pos] + new_base + seq[pos+1:]
    orig_codon_pos = (pos // 3) * 3
    orig_codon = transcribe(seq[orig_codon_pos:orig_codon_pos+3])
    mut_codon  = transcribe(mutated[orig_codon_pos:orig_codon_pos+3])
    orig_aa = CODON_TABLE.get(orig_codon,'???')
    mut_aa  = CODON_TABLE.get(mut_codon,'???')
    if orig_aa == mut_aa:
        impact = 'Synonymous'
        impact_color = '#4ade80'
    elif mut_aa == 'STOP':
        impact = 'Nonsense'
        impact_color = '#f87171'
    else:
        impact = 'Missense'
        impact_color = '#fbbf24'
    orig_protein = translate_seq(transcribe(seq))
    mut_protein  = translate_seq(transcribe(mutated))
    return {
        'original_seq': seq,
        'mutated_seq': mutated,
        'position': pos+1,
        'original_base': seq[pos],
        'new_base': new_base,
        'original_codon': orig_codon,
        'mutated_codon': mut_codon,
        'original_aa': orig_aa,
        'mutated_aa': mut_aa,
        'impact': impact,
        'impact_color': impact_color,
        'original_protein': ' - '.join(orig_protein),
        'mutated_protein': ' - '.join(mut_protein),
    }

# ── New: Primer design ────────────────────────────────────────────────────────
def design_primers(seq, primer_len=20):
    def primer_tm(p):
        a,t,g,c = p.count('A'),p.count('T'),p.count('G'),p.count('C')
        n = len(p)
        if n < 14: return round(2*(a+t)+4*(g+c),1)
        return round(81.5+0.41*((g+c)/n*100)-675/n,1)
    def primer_gc(p):
        return round((p.count('G')+p.count('C'))/len(p)*100,1)
    def hairpin_risk(p):
        rc = reverse_complement(p)
        for l in range(4, len(p)//2+1):
            if p[:l] in rc: return True
        return False

    fwd = seq[:primer_len]
    rev = reverse_complement(seq[-primer_len:])
    amplicon = len(seq)
    return {
        'forward': {'seq':fwd,'tm':primer_tm(fwd),'gc':primer_gc(fwd),'hairpin':hairpin_risk(fwd)},
        'reverse': {'seq':rev,'tm':primer_tm(rev),'gc':primer_gc(rev),'hairpin':hairpin_risk(rev)},
        'amplicon_size': amplicon,
        'primer_length': primer_len,
    }

# ── New: Restriction digest simulator ────────────────────────────────────────
def restriction_digest(seq, enzymes):
    cuts = []
    for enzyme in enzymes:
        site = RESTRICTION_ENZYMES.get(enzyme)
        if not site: continue
        for pos in find_motif_positions(seq, site):
            cuts.append(pos + len(site)//2)  # cut in middle of site
    cuts = sorted(set(cuts))
    if not cuts:
        return {'fragments':[],'cuts':[],'note':'No cut sites found for selected enzymes.'}
    boundaries = [0] + cuts + [len(seq)]
    fragments = []
    for i in range(len(boundaries)-1):
        size = boundaries[i+1] - boundaries[i]
        fragments.append({'start':boundaries[i]+1,'end':boundaries[i+1],'size':size})
    return {'fragments': sorted(fragments, key=lambda x:-x['size']), 'cuts': cuts}

# ── New: PCR predictor ────────────────────────────────────────────────────────
def pcr_predict(seq, fwd_primer, rev_primer):
    fwd = fwd_primer.upper()
    rev = rev_primer.upper()
    rev_rc = reverse_complement(rev)
    fwd_pos = seq.find(fwd)
    rev_pos = seq.find(rev_rc)
    if fwd_pos == -1 or rev_pos == -1 or fwd_pos >= rev_pos:
        return {'found': False, 'message': 'Primers not found in correct orientation.'}
    amplicon = seq[fwd_pos:rev_pos+len(rev_rc)]
    return {
        'found': True,
        'fwd_pos': fwd_pos+1,
        'rev_pos': rev_pos+len(rev_rc),
        'amplicon_size': len(amplicon),
        'amplicon_gc': gc_content(amplicon),
        'amplicon_seq': amplicon if len(amplicon) <= 200 else amplicon[:100]+'...'+amplicon[-100:],
    }

# ── Charts ────────────────────────────────────────────────────────────────────
def make_charts(seq, protein_list, theme='dark'):
    bg   = '#0f172a' if theme=='dark' else '#f8fafc'
    card = '#1e293b' if theme=='dark' else '#ffffff'
    text = '#e2e8f0' if theme=='dark' else '#1e293b'
    muted= '#94a3b8' if theme=='dark' else '#64748b'
    grid = '#334155' if theme=='dark' else '#e2e8f0'

    freq = {b: seq.count(b) for b in 'ATCG'}
    window = max(5, len(seq)//10)
    gc_vals, positions = [], []
    for i in range(len(seq)-window+1):
        gc_vals.append(gc_content(seq[i:i+window]))
        positions.append(i+1)

    rows = 2 if protein_list else 1
    fig, axes = plt.subplots(rows, 2, figsize=(12, 4*rows))
    if rows == 1: axes = [axes]
    fig.patch.set_facecolor(bg)

    def style_ax(ax):
        ax.set_facecolor(card)
        ax.tick_params(colors=muted)
        for spine in ax.spines.values(): spine.set_edgecolor(grid)

    # Row 1: nucleotide freq + GC window
    style_ax(axes[0][0])
    colors = ['#4ade80','#f87171','#38bdf8','#fbbf24']
    bars = axes[0][0].bar(list('ATCG'), [freq[b] for b in 'ATCG'], color=colors, edgecolor=bg, width=0.5)
    axes[0][0].set_title('Nucleotide Frequency', color=text, pad=8)
    axes[0][0].set_xlabel('Base', color=muted); axes[0][0].set_ylabel('Count', color=muted)
    for bar, val in zip(bars, [freq[b] for b in 'ATCG']):
        axes[0][0].text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.05, str(val),
                        ha='center', color=text, fontweight='bold', fontsize=9)

    style_ax(axes[0][1])
    axes[0][1].plot(positions, gc_vals, color='#a78bfa', linewidth=2)
    axes[0][1].axhline(50, color=grid, linestyle='--', alpha=0.7, label='50%')
    axes[0][1].fill_between(positions, gc_vals, alpha=0.15, color='#a78bfa')
    axes[0][1].set_title(f'GC Content Sliding Window ({window}bp)', color=text, pad=8)
    axes[0][1].set_xlabel('Position (bp)', color=muted); axes[0][1].set_ylabel('GC %', color=muted)
    axes[0][1].set_ylim(0,100)
    axes[0][1].legend(labelcolor=muted, facecolor=card, edgecolor=grid)

    # Row 2: protein hydrophobicity profile + AA composition
    if protein_list and rows == 2:
        style_ax(axes[1][0])
        hydro_vals = [AA_HYDRO.get(aa,0) for aa in protein_list]
        win2 = max(1, len(hydro_vals)//10)
        h_profile = [sum(hydro_vals[i:i+win2])/win2 for i in range(len(hydro_vals)-win2+1)]
        x = list(range(1, len(h_profile)+1))
        axes[1][0].bar(x, h_profile,
                       color=['#f87171' if v>0 else '#38bdf8' for v in h_profile], width=1.0)
        axes[1][0].axhline(0, color=grid, linewidth=1)
        axes[1][0].set_title('Protein Hydrophobicity Profile', color=text, pad=8)
        axes[1][0].set_xlabel('Residue', color=muted); axes[1][0].set_ylabel('Hydrophobicity', color=muted)
        red_p = mpatches.Patch(color='#f87171', label='Hydrophobic')
        blue_p = mpatches.Patch(color='#38bdf8', label='Hydrophilic')
        axes[1][0].legend(handles=[red_p,blue_p], labelcolor=muted, facecolor=card, edgecolor=grid)

        style_ax(axes[1][1])
        aa_counts = Counter(protein_list)
        aas = list(aa_counts.keys())
        cnts = [aa_counts[a] for a in aas]
        cmap = plt.cm.get_cmap('tab20', len(aas))
        axes[1][1].bar(aas, cnts, color=[cmap(i) for i in range(len(aas))], edgecolor=bg)
        axes[1][1].set_title('Amino Acid Composition', color=text, pad=8)
        axes[1][1].set_xlabel('Amino Acid', color=muted); axes[1][1].set_ylabel('Count', color=muted)
        axes[1][1].tick_params(axis='x', rotation=45)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor=bg)
    plt.close()
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()

def make_circular_plot(seq, features):
    """Simple circular genome / plasmid-style plot."""
    fig, ax = plt.subplots(figsize=(6,6))
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#0f172a')
    ax.set_aspect('equal'); ax.axis('off')
    n = len(seq)
    # Outer ring
    circle = plt.Circle((0,0), 1.0, fill=False, color='#334155', linewidth=8)
    ax.add_patch(circle)
    ax.text(0, 0, f'{n} bp', ha='center', va='center', color='#94a3b8', fontsize=11, fontweight='bold')
    # GC content arc coloring
    segs = min(360, n)
    seg_size = n / segs
    for i in range(segs):
        start_bp = int(i*seg_size)
        end_bp   = int((i+1)*seg_size)
        chunk = seq[start_bp:end_bp] if end_bp <= n else seq[start_bp:]
        if not chunk: continue
        gc = (chunk.count('G')+chunk.count('C'))/len(chunk)
        angle_start = 90 - i*(360/segs)
        angle_end   = 90 - (i+1)*(360/segs)
        color = plt.cm.RdYlGn(gc)
        theta = np.linspace(np.radians(angle_end), np.radians(angle_start), 5)
        r_in, r_out = 0.85, 1.0
        xs = np.concatenate([r_in*np.cos(theta), r_out*np.cos(theta[::-1])])
        ys = np.concatenate([r_in*np.sin(theta), r_out*np.sin(theta[::-1])])
        ax.fill(xs, ys, color=color, alpha=0.8)
    # Feature annotations
    feat_colors = {'ORF':'#6366f1','EcoRI':'#f87171','BamHI':'#fbbf24','HindIII':'#4ade80'}
    for feat in features[:8]:
        angle = 90 - (feat['pos']/n)*360
        rad = np.radians(angle)
        x, y = 1.15*np.cos(rad), 1.15*np.sin(rad)
        ax.plot([np.cos(rad), 1.08*np.cos(rad)],[np.sin(rad), 1.08*np.sin(rad)],
                color=feat_colors.get(feat['type'],'#94a3b8'), linewidth=1.5)
        ax.text(x, y, feat['label'], ha='center', va='center',
                color=feat_colors.get(feat['type'],'#94a3b8'), fontsize=6)
    ax.set_xlim(-1.5,1.5); ax.set_ylim(-1.5,1.5)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor='#0f172a')
    plt.close(); buf.seek(0)
    return base64.b64encode(buf.read()).decode()

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence',''))
        theme = data.get('theme','dark')
        rna = transcribe(seq)
        protein_list = translate_seq(rna, rna.find('AUG') if 'AUG' in rna else 0)
        tm, tm_rule = melting_temp(seq)
        orfs_6frame = find_all_orfs_6frame(seq)
        all_orfs = [o for f in orfs_6frame for o in f['orfs']]
        all_orfs_sorted = sorted(all_orfs, key=lambda x:-x['length'])
        props = protein_properties(protein_list)
        organism = data.get('organism','ecoli')

        # Circular plot features
        circ_features = []
        for o in all_orfs_sorted[:3]:
            circ_features.append({'pos':o['start'],'type':'ORF','label':f"ORF {o['length']}bp"})
        for enz, info in restriction_sites(seq).items():
            for p in info['positions'][:2]:
                circ_features.append({'pos':p,'type':enz,'label':enz})

        result = {
            'sequence': seq,
            'length': len(seq),
            'gc': gc_content(seq),
            'at': round(100-gc_content(seq),2),
            'freq': {b: seq.count(b) for b in 'ATCG'},
            'reverse_complement': reverse_complement(seq),
            'rna': rna,
            'protein': ' - '.join(protein_list) if protein_list else 'No start codon found.',
            'protein_list': protein_list,
            'tm': tm, 'tm_rule': tm_rule,
            'orfs_6frame': orfs_6frame,
            'top_orfs': all_orfs_sorted[:6],
            'aa_composition': [{'aa':aa,'count':n} for aa,n in sorted(Counter(protein_list).items())],
            'protein_props': props,
            'codon_usage': codon_usage(seq),
            'codon_bias': codon_bias_comparison(seq, organism),
            'restriction_sites': restriction_sites(seq),
            'palindromes': palindromes(seq),
            'motif_library': scan_motif_library(seq),
            'primers': design_primers(seq) if len(seq) >= 40 else None,
            'chart': make_charts(seq, protein_list, theme),
            'circular': make_circular_plot(seq, circ_features) if len(seq) >= 50 else None,
        }

        if data.get('motif'):
            m = data['motif'].upper().strip()
            result['motif'] = m
            result['motif_positions'] = find_motif_positions(seq, m)

        return jsonify(result)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'Analysis error: {str(e)}'}), 500

@app.route('/mutate', methods=['POST'])
def mutate():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence',''))
        result = simulate_mutation(seq, data.get('position',1), data.get('base','A'))
        return jsonify(result)
    except (ValueError, TypeError) as e:
        return jsonify({'error': str(e)}), 400

# ── Disease Links via multiple free APIs ─────────────────────────────────────
def query_disease(original_aa, mutated_aa, original_codon, mutated_codon, impact):
    import urllib.request, urllib.parse, json as _json

    result = {
        'queried': False, 'found': False,
        'disease': 'No known disease association in external database.',
        'severity': 'Unknown', 'severity_color': '#94a3b8',
        'source': '', 'variant_id': '', 'details': [],
        'clinvar_entries': [], 'omim_hint': ''
    }

    if impact == 'Synonymous':
        result['disease'] = 'Synonymous mutation — no amino acid change. Typically benign.'
        result['severity'] = 'Benign'
        result['severity_color'] = '#4ade80'
        return result

    if impact == 'Nonsense':
        result['queried'] = True
        result['disease'] = 'Nonsense mutation introduces a premature stop codon — likely disrupts protein function.'
        result['severity'] = 'Pathogenic'
        result['severity_color'] = '#f87171'
        result['details'].append('Premature stop codons are associated with loss-of-function diseases.')
        return result

    def safe_get(url):
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'CelestialCodon/1.0',
                'Accept': 'application/json'
            })
            with urllib.request.urlopen(req, timeout=6) as r:
                return _json.loads(r.read().decode())
        except Exception:
            return None

    result['queried'] = True
    sources_used = []
    diseases = []
    severities = []

    # ── API 1: MyVariant.info — search by amino acid ref/alt ──────────────
    mv_url = (f'https://myvariant.info/v1/query?'
              f'q=dbnsfp.aa.ref:{original_aa}+AND+dbnsfp.aa.alt:{mutated_aa}'
              f'&fields=clinvar,dbsnp,dbnsfp,cadd&size=5')
    mv_data = safe_get(mv_url)
    if mv_data and mv_data.get('hits'):
        sources_used.append('MyVariant.info')
        for hit in mv_data['hits']:
            if not result['variant_id']:
                result['variant_id'] = hit.get('_id', '')

            # ClinVar
            cv = hit.get('clinvar', {})
            if cv:
                rcv = cv.get('rcv', {})
                if isinstance(rcv, list): rcv = rcv[0] if rcv else {}
                sig = rcv.get('clinical_significance', '') or cv.get('clinical_significance', '')
                cond = rcv.get('conditions', {})
                if isinstance(cond, dict) and cond.get('name'):
                    diseases.append(cond['name'])
                elif isinstance(cond, list):
                    for c in cond:
                        if isinstance(c, dict) and c.get('name'):
                            diseases.append(c['name'])
                if sig: severities.append(str(sig))
                cv_id = cv.get('variant_id', '')
                if cv_id:
                    result['clinvar_entries'].append(f'ClinVar ID: {cv_id}')

            # dbSNP
            dbsnp = hit.get('dbsnp', {})
            if isinstance(dbsnp, dict) and dbsnp.get('rsid'):
                result['details'].append(f"dbSNP: {dbsnp['rsid']}")

            # CADD score (deleteriousness)
            cadd = hit.get('cadd', {})
            if isinstance(cadd, dict):
                phred = cadd.get('phred', '')
                if phred:
                    label = 'High' if float(phred) > 20 else 'Low'
                    result['details'].append(f'CADD score: {phred} ({label} deleteriousness)')

            # dbnsfp SIFT + PolyPhen
            dbnsfp = hit.get('dbnsfp', {})
            if isinstance(dbnsfp, dict):
                sift = dbnsfp.get('sift', {})
                if isinstance(sift, dict):
                    pred = str(sift.get('pred', ''))
                    if pred: result['details'].append(f'SIFT: {"Deleterious" if "D" in pred else "Tolerated"}')
                pp2 = dbnsfp.get('polyphen2', {})
                if isinstance(pp2, dict):
                    hdiv = pp2.get('hdiv', {})
                    if isinstance(hdiv, dict):
                        pred = str(hdiv.get('pred', ''))
                        if pred: result['details'].append(f'PolyPhen2: {"Damaging" if "D" in pred else "Benign"}')

    # ── API 2: NCBI ClinVar REST — search by protein change ───────────────
    clinvar_query = urllib.parse.quote(f'{original_aa}[Protein Change] AND {mutated_aa}[Protein Change]')
    cv_url = f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=clinvar&term={clinvar_query}&retmax=3&retmode=json'
    cv_data = safe_get(cv_url)
    if cv_data:
        ids = cv_data.get('esearchresult', {}).get('idlist', [])
        if ids:
            sources_used.append('NCBI ClinVar')
            # fetch summary for first ID
            sum_url = f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?db=clinvar&id={ids[0]}&retmode=json'
            sum_data = safe_get(sum_url)
            if sum_data:
                doc = sum_data.get('result', {}).get(ids[0], {})
                title = doc.get('title', '')
                sig = doc.get('clinical_significance', {})
                if isinstance(sig, dict):
                    desc = sig.get('description', '')
                    if desc: severities.append(desc)
                trait_set = doc.get('trait_set', [])
                if isinstance(trait_set, list):
                    for t in trait_set:
                        if isinstance(t, dict) and t.get('trait_name'):
                            diseases.append(t['trait_name'])
                if title and not diseases:
                    diseases.append(title)
                result['clinvar_entries'].append(f'ClinVar UID: {ids[0]}')

    # ── API 3: Ensembl Variant Effect Predictor (VEP) — by HGVS-like ─────
    # Search for known variants with this amino acid change in human
    ensembl_url = (f'https://rest.ensembl.org/variant_recoder/human/'
                   f'?content-type=application/json')  # placeholder — VEP needs real HGVS

    # ── Compile results ────────────────────────────────────────────────────
    if diseases:
        result['found'] = True
        result['disease'] = ' | '.join(list(dict.fromkeys(diseases))[:3])  # deduplicate
    else:
        result['disease'] = 'No known disease association found in queried databases.'

    sev_text = ' '.join(severities).lower()
    if 'pathogenic' in sev_text and 'benign' not in sev_text:
        result['severity'] = 'Pathogenic'
        result['severity_color'] = '#f87171'
    elif 'likely pathogenic' in sev_text:
        result['severity'] = 'Likely Pathogenic'
        result['severity_color'] = '#fb923c'
    elif 'benign' in sev_text or 'likely benign' in sev_text:
        result['severity'] = 'Benign'
        result['severity_color'] = '#4ade80'
    elif severities:
        result['severity'] = severities[0].title()
        result['severity_color'] = '#fbbf24'
    else:
        result['severity'] = 'Uncertain Significance'
        result['severity_color'] = '#fbbf24'

    # deduplicate details
    result['details'] = list(dict.fromkeys(result['details']))
    result['source'] = ' + '.join(sources_used) if sources_used else 'MyVariant.info + NCBI ClinVar'
    return result

@app.route('/disease', methods=['POST'])
def disease_lookup():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence', ''))
        position = data.get('position', 1)
        base = data.get('base', 'A')
        mut = simulate_mutation(seq, position, base)
        disease = query_disease(
            mut['original_aa'], mut['mutated_aa'],
            mut['original_codon'], mut['mutated_codon'],
            mut['impact']
        )
        return jsonify({**mut, 'disease_info': disease})
    except (ValueError, TypeError) as e:
        return jsonify({'error': str(e)}), 400

@app.route('/digest', methods=['POST'])
def digest():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence',''))
        enzymes = data.get('enzymes',[])
        return jsonify(restriction_digest(seq, enzymes))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/pcr', methods=['POST'])
def pcr():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence',''))
        fwd = data.get('fwd','').upper().strip()
        rev = data.get('rev','').upper().strip()
        if not fwd or not rev:
            return jsonify({'error':'Both primers required.'}), 400
        return jsonify(pcr_predict(seq, fwd, rev))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/export/fasta', methods=['POST'])
def export_fasta():
    from flask import Response
    data = request.get_json()
    seq = validate(data.get('sequence',''))
    name = data.get('name','sequence')
    fasta = f'>{name}\n'
    fasta += '\n'.join(seq[i:i+60] for i in range(0,len(seq),60))
    return Response(fasta, mimetype='text/plain',
                    headers={'Content-Disposition':f'attachment;filename={name}.fasta'})

@app.route('/export/xlsx', methods=['POST'])
def export_xlsx():
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    data = request.get_json()
    seq = validate(data.get('sequence',''))
    rna = transcribe(seq)
    protein_list = translate_seq(rna, rna.find('AUG') if 'AUG' in rna else 0)
    tm, tm_rule = melting_temp(seq)
    props = protein_properties(protein_list)
    orfs = find_all_orfs_6frame(seq)
    all_orfs = sorted([o for f in orfs for o in f['orfs']], key=lambda x: -x['length'])
    rsites = restriction_sites(seq)
    bias = codon_bias_comparison(seq, 'ecoli')
    aa_comp = Counter(protein_list)

    wb = Workbook()

    # ── Color palette ──
    C_TITLE   = PatternFill('solid', fgColor='1E3A5F')   # dark navy
    C_HEAD    = PatternFill('solid', fgColor='0EA5E9')   # sky blue
    C_HEAD2   = PatternFill('solid', fgColor='6366F1')   # indigo
    C_HEAD3   = PatternFill('solid', fgColor='059669')   # green
    C_HEAD4   = PatternFill('solid', fgColor='DC2626')   # red
    C_HEAD5   = PatternFill('solid', fgColor='D97706')   # amber
    C_HEAD6   = PatternFill('solid', fgColor='7C3AED')   # purple
    C_ROW_A   = PatternFill('solid', fgColor='EFF6FF')   # light blue row
    C_ROW_B   = PatternFill('solid', fgColor='FFFFFF')   # white row
    C_POS     = PatternFill('solid', fgColor='DCFCE7')   # light green (positive diff)
    C_NEG     = PatternFill('solid', fgColor='FEE2E2')   # light red (negative diff)
    C_AA      = PatternFill('solid', fgColor='F5F3FF')   # light purple

    F_WHITE   = Font(color='FFFFFF', bold=True, size=11)
    F_TITLE   = Font(color='FFFFFF', bold=True, size=14)
    F_BOLD    = Font(bold=True)
    F_MONO    = Font(name='Courier New', size=9)
    WRAP      = Alignment(wrap_text=False, vertical='top')
    WRAP_ON   = Alignment(wrap_text=True, vertical='top')
    CENTER    = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='CBD5E1')
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_title(ws, row, text, fill, cols=3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill
        c.font = Font(color='FFFFFF', bold=True, size=12)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 24

    def style_header(ws, row, headers, fill):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = fill; c.font = F_WHITE; c.alignment = CENTER; c.border = BORDER
            _track(ws, row, ci, h)

    def style_row(ws, row, values, fill, mono=False, wrap=False):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill; c.border = BORDER
            c.font = F_MONO if mono else Font(size=10)
            if wrap:
                c.alignment = WRAP_ON
                _track(ws, row, ci, v, wrap_col_width=60)
            else:
                c.alignment = WRAP
                _track(ws, row, ci, v)

    # ── Auto-fit tracker ──────────────────────────────────────────
    col_tracker = {}  # {ws_title: {col_index: max_char_len}}
    row_tracker = {}  # {ws_title: {row_index: num_lines}}

    def _track(ws, row, col, value, wrap_col_width=None):
        text = str(value) if value is not None else ''
        k = ws.title
        col_tracker.setdefault(k, {})
        row_tracker.setdefault(k, {})
        if wrap_col_width:
            lines = max(1, -(-len(text) // wrap_col_width))
            row_tracker[k][row] = max(row_tracker[k].get(row, 1), lines)
            col_tracker[k][col] = max(col_tracker[k].get(col, 0), min(len(text), wrap_col_width))
        else:
            col_tracker[k][col] = max(col_tracker[k].get(col, 0), len(text))
            row_tracker[k][row] = max(row_tracker[k].get(row, 1), 1)

    def apply_autofit(ws):
        """Autofit columns that don't have a manual width. Never shrink rows already set."""
        from openpyxl.utils import get_column_letter as gcl
        col_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    try:
                        col = gcl(cell.column)
                        col_widths[col] = max(col_widths.get(col, 0), len(str(cell.value)))
                    except Exception:
                        pass
        for col, max_len in col_widths.items():
            existing = ws.column_dimensions[col].width
            if not existing or existing == 0:
                ws.column_dimensions[col].width = min(max_len + 3, 60)
        # only set row heights that haven't been manually set (height == 0 means default)
        k = ws.title
        for row, lines in row_tracker.get(k, {}).items():
            current = ws.row_dimensions[row].height
            if not current or current == 0:
                ws.row_dimensions[row].height = max(16, lines * 15)

    def set_col_widths(ws, widths):
        pass  # replaced by autofit

    # ════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Summary'
    # Fixed column widths for summary — col B holds sequences so wrap at 80 chars
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 80
    ws1.column_dimensions['C'].width = 24

    style_title(ws1, 1, '🧬  CELESTIAL CODON — DNA Analysis Report', C_TITLE, 3)
    ws1.row_dimensions[2].height = 6  # spacer

    style_title(ws1, 3, 'Basic Statistics', C_HEAD, 3)
    style_header(ws1, 4, ['Property', 'Value', 'Unit'], C_HEAD)
    stats = [
        ('Sequence Length', len(seq), 'bp'),
        ('GC Content', gc_content(seq), '%'),
        ('AT Content', round(100-gc_content(seq),2), '%'),
        ('Melting Temperature', tm, f'deg C ({tm_rule})'),
        ('Nucleotide A', seq.count('A'), 'count'),
        ('Nucleotide T', seq.count('T'), 'count'),
        ('Nucleotide C', seq.count('C'), 'count'),
        ('Nucleotide G', seq.count('G'), 'count'),
    ]
    for i, row in enumerate(stats, 5):
        style_row(ws1, i, row, C_ROW_A if i%2==0 else C_ROW_B)

    ws1.row_dimensions[13].height = 6
    style_title(ws1, 14, 'Sequences', C_HEAD2, 3)
    style_header(ws1, 15, ['Type', 'Sequence', ''], C_HEAD2)
    seqs = [
        ('DNA Input', seq, ''),
        ('Reverse Complement', reverse_complement(seq), ''),
        ('mRNA Transcript', rna, ''),
        ('Protein', ' - '.join(protein_list) if protein_list else 'No start codon', ''),
    ]
    for i, row in enumerate(seqs, 16):
        style_row(ws1, i, row, C_ROW_A if i%2==0 else C_ROW_B, mono=True, wrap=True)
        seq_len = len(str(row[1]))
        lines = max(1, -(-seq_len // 80))
        ws1.row_dimensions[i].height = max(20, lines * 14)

    ws1.row_dimensions[20].height = 6
    style_title(ws1, 21, 'Protein Properties', C_HEAD3, 3)
    style_header(ws1, 22, ['Property', 'Value', 'Unit'], C_HEAD3)
    if props:
        prows = [
            ('Protein Length', props['length'], 'amino acids'),
            ('Molecular Weight', props['mw'], 'Da'),
            ('Isoelectric Point (pI)', props['pi'], ''),
            ('Avg Hydrophobicity', props['avg_hydrophobicity'], 'Kyte-Doolittle'),
            ('Classification', props['classification'], ''),
        ]
        for i, row in enumerate(prows, 23):
            style_row(ws1, i, row, C_ROW_A if i%2==0 else C_ROW_B)

    # ════════════════════════════════════════════
    # Sheet 2 — ORFs
    # ════════════════════════════════════════════
    ws2 = wb.create_sheet('ORFs')
    style_title(ws2, 1, 'Open Reading Frames — All 6 Frames', C_HEAD4, 6)
    style_header(ws2, 2, ['Strand', 'Frame', 'Start', 'End', 'Length (bp)', 'Sequence'], C_HEAD4)
    frame_fills = [
        PatternFill('solid', fgColor='EDE9FE'),
        PatternFill('solid', fgColor='E0E7FF'),
        PatternFill('solid', fgColor='DBEAFE'),
        PatternFill('solid', fgColor='FCE7F3'),
        PatternFill('solid', fgColor='FEF3C7'),
        PatternFill('solid', fgColor='DCFCE7'),
    ]
    frame_map = {('forward',1):0,('forward',2):1,('forward',3):2,
                 ('reverse',1):3,('reverse',2):4,('reverse',3):5}
    for ri, orf in enumerate(all_orfs, 3):
        fi = frame_map.get((orf['strand'], orf['frame']), 0)
        fill = frame_fills[fi]
        style_row(ws2, ri, [orf['strand'], orf['frame'], orf['start'],
                             orf['end'], orf['length'], orf['seq']], fill, mono=True)
    # ORF sheet: fixed widths, sequence col wraps so full sequence is visible
    for col, w in zip(['A','B','C','D','E'], [12, 10, 10, 10, 14]):
        ws2.column_dimensions[col].width = w
    ws2.column_dimensions['F'].width = 100
    for ri, orf in enumerate(all_orfs, 3):
        cell = ws2.cell(row=ri, column=6)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = F_MONO
        lines = max(1, -(-len(orf['seq']) // 100))
        ws2.row_dimensions[ri].height = max(16, lines * 13)
    # ════════════════════════════════════════════
    # Sheet 3 — Restriction Sites & Palindromes
    # ════════════════════════════════════════════
    ws3 = wb.create_sheet('Restriction & Palindromes')
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 40
    style_title(ws3, 1, 'Restriction Enzyme Sites', C_HEAD5, 3)
    style_header(ws3, 2, ['Enzyme', 'Recognition Site', 'Cut Positions'], C_HEAD5)
    if rsites:
        for i, (enzyme, info) in enumerate(rsites.items(), 3):
            fill = C_ROW_A if i%2==0 else C_ROW_B
            style_row(ws3, i, [enzyme, info['site'], ', '.join(map(str,info['positions']))], fill)
    else:
        ws3.cell(row=3, column=1, value='No restriction sites found')

    row_off = len(rsites)+5 if rsites else 5
    style_title(ws3, row_off, 'Palindromes / Hairpin Regions', C_HEAD6, 3)
    style_header(ws3, row_off+1, ['Start', 'End', 'Sequence'], C_HEAD6)
    for i, p in enumerate(palindromes(seq), row_off+2):
        fill = C_ROW_A if i%2==0 else C_ROW_B
        style_row(ws3, i, [p['start'], p['end'], p['seq']], fill, mono=True)

    # ════════════════════════════════════════════
    # Sheet 4 — Codon Bias
    # ════════════════════════════════════════════
    ws4 = wb.create_sheet('Codon Bias vs E.coli')
    ws4.column_dimensions['A'].width = 10
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 16
    ws4.column_dimensions['D'].width = 16
    ws4.column_dimensions['E'].width = 14
    style_title(ws4, 1, 'Codon Usage — Sequence vs E. coli K12 (per 1000 codons)', C_HEAD2, 5)
    style_header(ws4, 2, ['Codon', 'Amino Acid', 'Seq Freq', 'E.coli Freq', 'Difference'], C_HEAD2)
    for i, row in enumerate(bias, 3):
        fill = C_POS if row['diff'] > 0 else (C_NEG if row['diff'] < 0 else C_ROW_B)
        style_row(ws4, i, [row['codon'], row['aa'], row['seq_freq'], row['ref_freq'], row['diff']], fill, mono=True)

    # ════════════════════════════════════════════
    # Sheet 5 — Amino Acid Composition
    # ════════════════════════════════════════════
    ws5 = wb.create_sheet('Amino Acids')
    style_title(ws5, 1, 'Amino Acid Composition', C_HEAD3, 2)
    style_header(ws5, 2, ['Amino Acid', 'Count'], C_HEAD3)
    for i, (aa, count) in enumerate(sorted(aa_comp.items()), 3):
        style_row(ws5, i, [aa, count], C_AA)
    ws5.column_dimensions['A'].width = 20
    ws5.column_dimensions['B'].width = 12

    # ════════════════════════════════════════════
    # Sheet 6 — Nucleotide Table
    # ════════════════════════════════════════════
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint

    ws6 = wb.create_sheet('Nucleotide Table')
    ws6.merge_cells('A1:B1')
    c = ws6.cell(row=1, column=1, value='Nucleotide Composition — Table')
    c.fill = C_HEAD
    c.font = Font(color='FFFFFF', bold=True, size=12)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws6.row_dimensions[1].height = 22
    style_header(ws6, 2, ['Nucleotide', 'Count'], C_HEAD)
    nuc_data = [('A', seq.count('A')), ('T', seq.count('T')),
                ('C', seq.count('C')), ('G', seq.count('G'))]
    for i, (base, count) in enumerate(nuc_data, 3):
        style_row(ws6, i, [base, count], C_ROW_A if i%2==0 else C_ROW_B)
    # fixed widths for clean table view
    ws6.column_dimensions['A'].width = 14
    ws6.column_dimensions['B'].width = 10

    # ════════════════════════════════════════════
    # Sheet 7 — Pie Chart (nucleotide composition)
    # ════════════════════════════════════════════
    ws7 = wb.create_sheet('Charts')
    ws7.merge_cells('A1:E1')
    c7 = ws7.cell(row=1, column=1, value='Nucleotide Composition — Pie Chart')
    c7.fill = C_HEAD
    c7.font = Font(color='FFFFFF', bold=True, size=12)
    c7.alignment = Alignment(horizontal='center', vertical='center')
    ws7.row_dimensions[1].height = 24

    pie = PieChart()
    pie.title = 'Nucleotide Composition'
    pie.style = 10
    labels = Reference(ws6, min_col=1, min_row=3, max_row=6)
    data   = Reference(ws6, min_col=2, min_row=2, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 20; pie.height = 12
    # color slices: A=green, T=red, C=blue, G=amber
    slice_colors = ['4ADE80','F87171','38BDF8','FBBF24']
    for idx, color in enumerate(slice_colors):
        pt = DataPoint(idx=idx, explosion=1)
        from openpyxl.drawing.fill import PatternFillProperties
        from openpyxl.utils.units import pixels_to_EMU
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)
    ws7.add_chart(pie, 'A3')

    # Apply auto-fit to all sheets
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
        apply_autofit(ws)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment;filename=dna_report.xlsx'})

@app.route('/export/csv', methods=['POST'])
def export_csv():
    from flask import Response
    import csv, io as _io
    data = request.get_json()
    seq = validate(data.get('sequence',''))
    rna = transcribe(seq)
    protein_list = translate_seq(rna, rna.find('AUG') if 'AUG' in rna else 0)
    tm, tm_rule = melting_temp(seq)
    props = protein_properties(protein_list)
    orfs = find_all_orfs_6frame(seq)
    all_orfs = sorted([o for f in orfs for o in f['orfs']], key=lambda x: -x['length'])
    rsites = restriction_sites(seq)
    bias = codon_bias_comparison(seq, 'ecoli')
    aa_comp = Counter(protein_list)

    buf = _io.StringIO()
    w = csv.writer(buf)

    # Section 1: Summary
    w.writerow(['=== CELESTIAL CODON - DNA ANALYSIS REPORT ===', '', ''])
    w.writerow(['Generated for sequence length:', f'{len(seq)} bp', ''])
    w.writerow([])
    w.writerow(['--- BASIC STATISTICS ---', '', ''])
    w.writerow(['Property', 'Value', 'Unit'])
    w.writerow(['Sequence Length', len(seq), 'bp'])
    w.writerow(['GC Content', gc_content(seq), '%'])
    w.writerow(['AT Content', round(100 - gc_content(seq), 2), '%'])
    w.writerow(['Melting Temperature', tm, f'deg C ({tm_rule})'])
    w.writerow(['Nucleotide A', seq.count('A'), 'count'])
    w.writerow(['Nucleotide T', seq.count('T'), 'count'])
    w.writerow(['Nucleotide C', seq.count('C'), 'count'])
    w.writerow(['Nucleotide G', seq.count('G'), 'count'])
    w.writerow([])

    # Section 2: Sequences
    w.writerow(['--- SEQUENCES ---', '', ''])
    w.writerow(['Type', 'Sequence', ''])
    w.writerow(['DNA (Input)', seq, ''])
    w.writerow(['Reverse Complement', reverse_complement(seq), ''])
    w.writerow(['mRNA Transcript', rna, ''])
    w.writerow(['Protein', ' - '.join(protein_list) if protein_list else 'No start codon', ''])
    w.writerow([])

    # Section 3: Protein properties
    w.writerow(['--- PROTEIN PROPERTIES ---', '', ''])
    w.writerow(['Property', 'Value', 'Unit'])
    if props:
        w.writerow(['Protein Length', props['length'], 'aa'])
        w.writerow(['Molecular Weight', props['mw'], 'Da'])
        w.writerow(['Isoelectric Point (pI)', props['pi'], ''])
        w.writerow(['Avg Hydrophobicity', props['avg_hydrophobicity'], '(Kyte-Doolittle)'])
        w.writerow(['Classification', props['classification'], ''])
    else:
        w.writerow(['No protein translated', '', ''])
    w.writerow([])

    # Section 4: Amino acid composition
    if aa_comp:
        w.writerow(['--- AMINO ACID COMPOSITION ---', '', ''])
        w.writerow(['Amino Acid', 'Count', ''])
        for aa, count in sorted(aa_comp.items()):
            w.writerow([aa, count, ''])
        w.writerow([])

    # Section 5: ORFs
    w.writerow(['--- OPEN READING FRAMES (Top 6) ---', '', ''])
    w.writerow(['Strand', 'Frame', 'Start', 'End', 'Length (bp)', 'Sequence'])
    for orf in all_orfs[:6]:
        w.writerow([orf['strand'], orf['frame'], orf['start'], orf['end'], orf['length'], orf['seq']])
    w.writerow([])

    # Section 6: Restriction sites
    w.writerow(['--- RESTRICTION SITES ---', '', ''])
    w.writerow(['Enzyme', 'Recognition Site', 'Positions'])
    if rsites:
        for enzyme, info in rsites.items():
            w.writerow([enzyme, info['site'], ', '.join(map(str, info['positions']))])
    else:
        w.writerow(['None found', '', ''])
    w.writerow([])

    # Section 7: Codon usage vs E.coli
    w.writerow(['--- CODON USAGE vs E.COLI (per 1000 codons) ---', '', ''])
    w.writerow(['Codon', 'Amino Acid', 'Sequence Freq', 'E.coli Ref Freq', 'Difference'])
    for row in bias:
        w.writerow([row['codon'], row['aa'], row['seq_freq'], row['ref_freq'], row['diff']])

    csv_bytes = '\ufeff' + buf.getvalue()  # BOM for Excel UTF-8 auto-detection
    return Response(csv_bytes.encode('utf-8'),
                    mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': 'attachment;filename=celestial_codon_report.csv'})

# ── New: CpG Island Detection ─────────────────────────────────────────────────
def find_cpg_islands(seq, window=200, step=1, gc_thresh=50, cpg_thresh=0.6):
    islands = []
    in_island = False
    island_start = 0
    for i in range(0, len(seq) - window + 1, step):
        w = seq[i:i+window]
        gc = gc_content(w)
        cpg_obs = w.count('CG')
        c_count = w.count('C'); g_count = w.count('G')
        cpg_exp = (c_count * g_count) / window if window > 0 else 0
        ratio = cpg_obs / cpg_exp if cpg_exp > 0 else 0
        is_island = gc >= gc_thresh and ratio >= cpg_thresh
        if is_island and not in_island:
            island_start = i + 1; in_island = True
        elif not is_island and in_island:
            islands.append({'start': island_start, 'end': i+window, 'length': i+window-island_start})
            in_island = False
    if in_island:
        islands.append({'start': island_start, 'end': len(seq), 'length': len(seq)-island_start})
    return islands[:10]

# ── New: Secondary Structure Prediction (Chou-Fasman) ────────────────────────
CF_HELIX = {'Ala':1.42,'Arg':0.98,'Asn':0.67,'Asp':1.01,'Cys':0.70,'Gln':1.11,
            'Glu':1.51,'Gly':0.57,'His':1.00,'Ile':1.08,'Leu':1.21,'Lys':1.16,
            'Met':1.45,'Phe':1.13,'Pro':0.57,'Ser':0.77,'Thr':0.83,'Trp':1.08,
            'Tyr':0.69,'Val':1.06}
CF_SHEET = {'Ala':0.83,'Arg':0.93,'Asn':0.89,'Asp':0.54,'Cys':1.19,'Gln':1.10,
            'Glu':0.37,'Gly':0.75,'His':0.87,'Ile':1.60,'Leu':1.30,'Lys':0.74,
            'Met':1.05,'Phe':1.38,'Pro':0.55,'Ser':0.75,'Thr':1.19,'Trp':1.37,
            'Tyr':1.47,'Val':1.70}

def predict_secondary_structure(protein_list, window=6):
    if not protein_list: return []
    result = []
    for i in range(len(protein_list)):
        w = protein_list[max(0,i-window//2):i+window//2+1]
        h = sum(CF_HELIX.get(aa,1.0) for aa in w)/len(w)
        s = sum(CF_SHEET.get(aa,1.0) for aa in w)/len(w)
        struct = 'H' if h>1.03 and h>s else ('E' if s>1.05 and s>h else 'C')
        result.append({'aa':protein_list[i],'pos':i+1,'struct':struct,
                       'helix_score':round(h,3),'sheet_score':round(s,3)})
    return result

# ── New: ELI5 Explanations ────────────────────────────────────────────────────
def eli5_explain(data):
    seq = data.get('sequence','')
    gc = data.get('gc',0); tm = data.get('tm',0)
    protein = data.get('protein','')
    orfs = data.get('top_orfs',[]); props = data.get('protein_props') or {}
    lines = [
        f"Your DNA sequence is {len(seq)} letters long — like a sentence written in a 4-letter alphabet: A, T, C, and G.",
        f"About {gc}% of your sequence is G and C. These letters stick together more strongly, like a tighter handshake.",
        f"If you heated this DNA in a lab, it would start to fall apart at around {tm}°C — that's called the melting temperature.",
    ]
    if 'AUG' in seq.replace('T','U'):
        lines.append("Your sequence has a start signal (ATG) — like a 'begin reading here' marker that tells the cell where to start making a protein.")
    if protein and 'No start' not in protein:
        aa_count = len(protein.split(' - '))
        lines.append(f"The protein made from this DNA is {aa_count} amino acids long. Amino acids are the building blocks of proteins, like beads on a necklace.")
    if props.get('classification') == 'Hydrophobic':
        lines.append("The protein avoids water (hydrophobic). These proteins often live inside cell membranes.")
    elif props.get('classification') == 'Hydrophilic':
        lines.append("The protein likes water (hydrophilic). These proteins usually float in the watery parts of the cell.")
    if props.get('pi'):
        pi = props['pi']
        lines.append(f"The protein's pI is {pi} — {'it carries a negative charge' if pi<7 else 'it carries a positive charge'} at normal body pH (7.4).")
    if orfs:
        lines.append(f"We found {len(orfs)} possible protein-coding regions (ORFs). The longest is {orfs[0]['length']} letters long.")
    return lines

# ── New: Quiz Generator ───────────────────────────────────────────────────────
def generate_quiz(seq, protein_list, all_orfs, rsites, gc):
    import random
    questions = [
        {'q': f'What is the GC content of this sequence?',
         'options': [f'{gc}%', f'{round(gc+10,1)}%', f'{round(gc-10,1)}%', f'{round(gc+5,1)}%'],
         'answer': 0, 'explanation': f'GC content = (G+C bases / total bases) × 100 = {gc}%'},
        {'q': 'What does the start codon ATG code for?',
         'options': ['Methionine (Met)','Alanine (Ala)','Glycine (Gly)','Stop signal'],
         'answer': 0, 'explanation': 'ATG always codes for Methionine and signals the start of translation.'},
        {'q': f'How many base pairs does this sequence have?',
         'options': [str(len(seq)), str(len(seq)+5), str(len(seq)-3), str(len(seq)*2)],
         'answer': 0, 'explanation': f'The sequence contains {len(seq)} nucleotides.'},
        {'q': f'What is the reverse complement of 5\'-{seq[:6]}-3\'?',
         'options': [reverse_complement(seq[:6]), seq[:6], transcribe(seq[:6]), seq[:6][::-1]],
         'answer': 0, 'explanation': f'Reverse complement: flip and swap A↔T, C↔G. Result: {reverse_complement(seq[:6])}'},
        {'q': 'What type of mutation changes one amino acid to a different one?',
         'options': ['Missense','Nonsense','Synonymous','Frameshift'],
         'answer': 0, 'explanation': 'A missense mutation changes the codon so a different amino acid is inserted.'},
    ]
    if rsites:
        enzyme = list(rsites.keys())[0]; site = rsites[enzyme]['site']
        questions.append({'q': f'Which enzyme recognizes the sequence {site}?',
                          'options': [enzyme,'EcoRV','TaqI','HaeIII'],
                          'answer': 0, 'explanation': f'{enzyme} recognizes and cuts at {site}.'})
    random.shuffle(questions)
    for q in questions:
        correct = q['options'][q['answer']]
        random.shuffle(q['options'])
        q['answer'] = q['options'].index(correct)
    return questions[:5]

# ── New Routes ────────────────────────────────────────────────────────────────
@app.route('/eli5', methods=['POST'])
def eli5():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence',''))
        rna = transcribe(seq)
        protein_list = translate_seq(rna, rna.find('AUG') if 'AUG' in rna else 0)
        props = protein_properties(protein_list)
        orfs = sorted([o for f in find_all_orfs_6frame(seq) for o in f['orfs']], key=lambda x:-x['length'])
        payload = {'sequence':seq,'gc':gc_content(seq),'tm':melting_temp(seq)[0],
                   'protein':' - '.join(protein_list),'top_orfs':orfs[:3],'protein_props':props}
        return jsonify({'lines': eli5_explain(payload)})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/secondary', methods=['POST'])
def secondary():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence',''))
        rna = transcribe(seq)
        protein_list = translate_seq(rna, rna.find('AUG') if 'AUG' in rna else 0)
        return jsonify({'secondary_structure': predict_secondary_structure(protein_list),
                        'cpg_islands': find_cpg_islands(seq)})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/quiz', methods=['POST'])
def quiz():
    try:
        data = request.get_json()
        seq = validate(data.get('sequence',''))
        rna = transcribe(seq)
        protein_list = translate_seq(rna, rna.find('AUG') if 'AUG' in rna else 0)
        orfs = sorted([o for f in find_all_orfs_6frame(seq) for o in f['orfs']], key=lambda x:-x['length'])
        return jsonify({'questions': generate_quiz(seq, protein_list, orfs, restriction_sites(seq), gc_content(seq))})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/export/json', methods=['POST'])
def export_json():
    from flask import Response
    import json as _json
    data = request.get_json()
    seq = validate(data.get('sequence',''))
    rna = transcribe(seq)
    protein_list = translate_seq(rna, rna.find('AUG') if 'AUG' in rna else 0)
    tm, tm_rule = melting_temp(seq)
    props = protein_properties(protein_list)
    orfs = sorted([o for f in find_all_orfs_6frame(seq) for o in f['orfs']], key=lambda x:-x['length'])
    report = {
        'meta': {'tool':'Celestial Codon','sequence_length':len(seq)},
        'basic_stats': {'gc_content':gc_content(seq),'at_content':round(100-gc_content(seq),2),
                        'melting_temp':tm,'melting_temp_rule':tm_rule,
                        'nucleotide_freq':{b:seq.count(b) for b in 'ATCG'}},
        'sequences': {'dna':seq,'rna':rna,'reverse_complement':reverse_complement(seq),
                      'protein':' - '.join(protein_list) if protein_list else 'No start codon'},
        'protein_properties': props,
        'orfs': orfs[:10],
        'restriction_sites': restriction_sites(seq),
        'palindromes': palindromes(seq),
        'cpg_islands': find_cpg_islands(seq),
        'secondary_structure': predict_secondary_structure(protein_list),
        'codon_usage': codon_usage(seq),
        'motif_library': scan_motif_library(seq),
        'primers': design_primers(seq) if len(seq)>=40 else None,
    }
    return Response(_json.dumps(report, indent=2), mimetype='application/json',
                    headers={'Content-Disposition':'attachment;filename=celestial_codon_report.json'})

# ── Gemini route (DISABLED — requires subscription / quota exceeded) ──────────
# @app.route('/ask_ai', methods=['POST'])
# @app.route('/ask-ai', methods=['POST'])
# def ask_ai():
#     """Gemini 2.0 Flash — requires valid API key with available quota.
#     Free tier is very limited. Switch to Groq for free usage."""
#     data = request.get_json()
#     question = data.get('question', '')
#     sequence = data.get('sequence', '')
#     prompt = f"You are a bioinformatics AI assistant.\nDNA Sequence:\n{sequence}\nQuestion:\n{question}\nGive a clear scientific answer."
#     try:
#         response = gemini_client.models.generate_content(model='gemini-2.0-flash', contents=prompt)
#         return jsonify({'answer': response.text})
#     except Exception as e:
#         err_str = str(e)
#         if '429' in err_str or 'RESOURCE_EXHAUSTED' in err_str:
#             return jsonify({'answer': None, 'quota_exceeded': True, 'retry_after': 35})
#         return jsonify({'answer': f'Error: {err_str}'})


# ── Ollama route (DISABLED — requires Ollama running locally) ─────────────────
# def ask_ollama(prompt):
#     """Ollama — runs locally only. Install from https://ollama.com
#     then run: ollama serve && ollama pull llama3.2"""
#     import requests as _req
#     resp = _req.post('http://localhost:11434/api/generate',
#                      json={'model': 'llama3.2', 'prompt': prompt, 'stream': False}, timeout=60)
#     resp.raise_for_status()
#     return resp.json()['response']
#
# @app.route('/ask_local_ai', methods=['POST'])
# def ask_local_ai():
#     """Ollama local AI — requires Ollama running locally (can run offline)."""
#     try:
#         data = request.get_json()
#         question = data.get('question', '').strip()
#         sequence = data.get('sequence', '').strip()
#         if not question:
#             return jsonify({'error': 'No question provided'}), 400
#         prompt = f"You are a bioinformatics AI assistant.\nDNA Sequence:\n{sequence or 'No sequence provided'}\nQuestion:\n{question}\nGive a clear, concise scientific answer."
#         answer = ask_ollama(prompt)
#         return jsonify({'answer': answer})
#     except Exception as e:
#         err = str(e)
#         if 'Connection refused' in err or 'Failed to establish' in err:
#             return jsonify({'error': 'Ollama is not running. Start it with: ollama serve'}), 503
#         return jsonify({'error': err}), 500


# ── Groq route (ACTIVE — free tier available) ─────────────────────────────────
@app.route('/ask_groq', methods=['POST'])
def ask_groq():
    """Groq API — free tier with generous limits. Get your key at https://console.groq.com"""
    try:
        if not _groq_available:
            return jsonify({'answer': 'Groq package not installed. Run: pip install groq'})

        api_key = os.getenv("GROQ_API_KEY", "")
        if not api_key:
            return jsonify({'answer': 'GROQ_API_KEY not set. Get a free key at https://console.groq.com'})

        data = request.get_json(silent=True) or {}
        question = data.get('question', '').strip()
        sequence = data.get('sequence', '').strip()

        if not question:
            return jsonify({'answer': 'No question provided.'})

        prompt = f"""You are Armaani, a helpful AI assistant with expertise in bioinformatics and molecular biology.

{f'DNA Sequence context: {sequence}' if sequence else ''}

User: {question}

Answer concisely and directly. For biology/DNA questions use your expertise. For general questions just answer normally."""

        groq = get_groq_client()
        response = groq.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=512,
        )
        return jsonify({'answer': response.choices[0].message.content})

    except Exception as e:
        err = str(e)
        if '429' in err or 'rate_limit' in err.lower():
            return jsonify({'answer': 'Groq rate limit reached. Please wait a moment and try again.'})
        return jsonify({'answer': f'Groq error: {err}'})


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5001))
    app.run(debug=False, use_reloader=False, host='0.0.0.0', port=port)
